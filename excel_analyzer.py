"""
Excel Analyzer Module

This module contains core functions for analyzing Excel workbooks containing
Source-Target mapping data. It handles extracting metadata, identifying column
structures, and parsing mapping data from worksheets.
"""

import logging
from typing import Dict, List, Optional, Tuple, Any
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from data_models import (
    TabMetadata, ColumnMapping, TabAnalysis, MappingRecord
)
from config import (
    SYSTEM_NAMES_ROW, HEADERS_ROW, DATA_START_ROW,
    SOURCE_SYSTEM_COLUMN, DEFAULT_TARGET_SYSTEM_COLUMN,
    COLUMN_NAME_MAPPINGS, STANDARD_COLUMN_ORDER,
    SYSTEM_NAME_MAX_SEARCH_COLUMNS, SYSTEM_NAME_MIN_LENGTH,
    MIN_COLUMN_HEADER_LENGTH, MAX_EMPTY_COLUMNS_BETWEEN_SECTIONS,
    CASE_SENSITIVE_COMPARISON, TRIM_WHITESPACE, IGNORE_EMPTY_CELLS,
    MIN_MAPPING_FIELDS, SKIP_HIDDEN_TABS, PROCESS_HIDDEN_TABS
)

logger = logging.getLogger(__name__)


def extract_tab_metadata(worksheet: Worksheet) -> TabMetadata:
    """
    Extract metadata from a worksheet tab including system names.
    
    Args:
        worksheet: The openpyxl worksheet to analyze
        
    Returns:
        TabMetadata object with extracted information
    """
    metadata = TabMetadata()
    metadata.tab_name = worksheet.title
    metadata.max_row = worksheet.max_row
    metadata.max_column = worksheet.max_column
    
    try:
        # Extract source system name from column A, row 9
        source_cell = worksheet.cell(SYSTEM_NAMES_ROW, SOURCE_SYSTEM_COLUMN)
        metadata.source_system = str(source_cell.value or "").strip()
        
        # Find target system name - first non-empty cell after A9 in row 9
        target_system = ""
        target_column = SOURCE_SYSTEM_COLUMN + 1
        
        for col in range(SOURCE_SYSTEM_COLUMN + 1, min(SYSTEM_NAME_MAX_SEARCH_COLUMNS + 1, metadata.max_column + 1)):
            cell = worksheet.cell(SYSTEM_NAMES_ROW, col)
            cell_value = str(cell.value or "").strip()
            if cell_value and len(cell_value) >= SYSTEM_NAME_MIN_LENGTH:
                target_system = cell_value
                target_column = col
                break
        
        metadata.target_system = target_system
        metadata.target_system_column = target_column
        
        logger.debug(f"Tab '{metadata.tab_name}': Source='{metadata.source_system}' (Col {SOURCE_SYSTEM_COLUMN}), "
                    f"Target='{metadata.target_system}' (Col {target_column})")
        
    except Exception as e:
        logger.error(f"Error extracting metadata from tab '{worksheet.title}': {e}")
        
    return metadata


def normalize_column_name(column_name: str) -> str:
    """
    Normalize a column name for consistent comparison.
    
    Args:
        column_name: The raw column name from Excel
        
    Returns:
        Normalized column name or empty string if not recognized
    """
    if not column_name:
        return ""
        
    # Clean the column name
    normalized = str(column_name).lower().strip()
    if not CASE_SENSITIVE_COMPARISON:
        normalized = normalized.lower()
    if TRIM_WHITESPACE:
        normalized = normalized.strip()
    
    # Find matching standard column type
    for standard_name, variations in COLUMN_NAME_MAPPINGS.items():
        if normalized in [v.lower() for v in variations]:
            return standard_name
    
    return ""


def identify_column_structure(worksheet: Worksheet, metadata: TabMetadata) -> ColumnMapping:
    """
    Identify the column structure by analyzing row 10 headers.
    
    Args:
        worksheet: The openpyxl worksheet to analyze
        metadata: Tab metadata containing system information
        
    Returns:
        ColumnMapping object with source and target column positions
    """
    column_mapping = ColumnMapping()
    
    try:
        # Extract all headers from row 10
        headers = {}
        for col in range(1, metadata.max_column + 1):
            cell = worksheet.cell(HEADERS_ROW, col)
            header_value = str(cell.value or "").strip()
            if header_value and len(header_value) >= MIN_COLUMN_HEADER_LENGTH:
                headers[col] = header_value
        
        column_mapping.all_headers = headers
        logger.debug(f"Found {len(headers)} headers in tab '{metadata.tab_name}'")
        
        # Identify source and target sections
        source_columns, target_columns = _detect_source_target_sections(headers, metadata)
        
        # Map standard column names to positions
        for col_num, header in headers.items():
            normalized = normalize_column_name(header)
            if normalized:
                if col_num in source_columns:
                    # For source columns, only keep first occurrence
                    if normalized not in column_mapping.source_columns:
                        column_mapping.source_columns[normalized] = col_num
                elif col_num in target_columns:
                    # For target columns, prioritize certain column names for duplicates
                    if normalized not in column_mapping.target_columns:
                        column_mapping.target_columns[normalized] = col_num
                    else:
                        # Handle duplicate target columns - prioritize Comments over Description
                        existing_col = column_mapping.target_columns[normalized]
                        existing_header = headers[existing_col].lower()
                        current_header = header.lower()
                        
                        # Prioritization rules for description fields
                        if normalized == 'description':
                            if 'comment' in current_header and 'comment' not in existing_header:
                                # Current header contains "comment", prefer it
                                column_mapping.target_columns[normalized] = col_num
                                logger.debug(f"Prioritizing '{header}' over '{headers[existing_col]}' for description field")
                        # Add more prioritization rules here if needed
        
        logger.debug(f"Tab '{metadata.tab_name}': Source cols={len(column_mapping.source_columns)}, "
                    f"Target cols={len(column_mapping.target_columns)}")
        
    except Exception as e:
        logger.error(f"Error identifying column structure in tab '{metadata.tab_name}': {e}")
    
    return column_mapping


def _detect_source_target_sections(headers: Dict[int, str], metadata: TabMetadata) -> Tuple[List[int], List[int]]:
    """
    Detect which columns belong to source vs target sections.
    
    Args:
        headers: Dictionary of column number to header text
        metadata: Tab metadata with target system column info
        
    Returns:
        Tuple of (source_column_list, target_column_list)
    """
    source_columns = []
    target_columns = []
    
    # Find the approximate boundary between source and target
    # Usually target section starts around the target system column
    boundary_col = metadata.target_system_column
    
    # Look for duplicate column names to identify sections
    header_counts = {}
    for col_num, header in headers.items():
        normalized = normalize_column_name(header)
        if normalized:
            if normalized not in header_counts:
                header_counts[normalized] = []
            header_counts[normalized].append(col_num)
    
    # If we find duplicate column types, assign first occurrence to source, second to target
    assigned_target = set()
    for normalized_name, col_list in header_counts.items():
        if len(col_list) >= 2:
            # Sort by column number
            col_list.sort()
            source_columns.extend(col_list[:1])  # First occurrence -> source
            target_columns.extend(col_list[1:])  # Later occurrences -> target
            assigned_target.update(col_list[1:])
    
    # For remaining columns, use boundary-based assignment
    for col_num, header in headers.items():
        if col_num not in assigned_target and col_num not in source_columns:
            if col_num < boundary_col:
                source_columns.append(col_num)
            else:
                # Check if there are too many empty columns before this one
                gap_size = _count_empty_columns_before(col_num, headers, boundary_col)
                if gap_size <= MAX_EMPTY_COLUMNS_BETWEEN_SECTIONS:
                    target_columns.append(col_num)
                else:
                    source_columns.append(col_num)
    
    return sorted(source_columns), sorted(target_columns)


def _count_empty_columns_before(col_num: int, headers: Dict[int, str], start_col: int) -> int:
    """Count empty columns between start_col and col_num."""
    count = 0
    for c in range(start_col, col_num):
        if c not in headers:
            count += 1
    return count


def extract_mappings_from_tab(worksheet: Worksheet, metadata: TabMetadata, 
                            column_mapping: ColumnMapping) -> List[MappingRecord]:
    """
    Extract all mapping records from a worksheet tab.
    
    Args:
        worksheet: The openpyxl worksheet to analyze
        metadata: Tab metadata
        column_mapping: Column structure information
        
    Returns:
        List of MappingRecord objects
    """
    mappings = []
    
    try:
        # Process each data row starting from row 11
        for row_num in range(DATA_START_ROW, metadata.max_row + 1):
            mapping = _extract_single_mapping(worksheet, row_num, column_mapping)
            if mapping and mapping.is_valid():
                mapping.row_number = row_num
                mappings.append(mapping)
        
        logger.debug(f"Extracted {len(mappings)} valid mappings from tab '{metadata.tab_name}'")
        
    except Exception as e:
        logger.error(f"Error extracting mappings from tab '{metadata.tab_name}': {e}")
    
    return mappings


def _extract_single_mapping(worksheet: Worksheet, row_num: int, 
                          column_mapping: ColumnMapping) -> Optional[MappingRecord]:
    """
    Extract a single mapping record from a worksheet row.
    
    Args:
        worksheet: The openpyxl worksheet
        row_num: Row number to extract from
        column_mapping: Column structure information
        
    Returns:
        MappingRecord object or None if row is empty/invalid
    """
    mapping = MappingRecord()
    
    try:
        # Extract source fields
        source_canonical_col = column_mapping.get_source_column('canonical_name')
        source_field_col = column_mapping.get_source_column('field')
        
        if source_canonical_col:
            cell = worksheet.cell(row_num, source_canonical_col)
            mapping.source_canonical = _clean_cell_value(cell.value)
        
        if source_field_col:
            cell = worksheet.cell(row_num, source_field_col)
            mapping.source_field = _clean_cell_value(cell.value)
        
        # Extract target fields
        target_canonical_col = column_mapping.get_target_column('canonical_name')
        target_field_col = column_mapping.get_target_column('field')
        
        if target_canonical_col:
            cell = worksheet.cell(row_num, target_canonical_col)
            mapping.target_canonical = _clean_cell_value(cell.value)
        
        if target_field_col:
            cell = worksheet.cell(row_num, target_field_col)
            mapping.target_field = _clean_cell_value(cell.value)
        
        # Extract all other fields for comparison using original column names
        all_fields = {}
        
        # Get all headers for original names
        all_headers = column_mapping.all_headers
        
        # Source fields - use original column names
        for field_type, col_num in column_mapping.source_columns.items():
            cell = worksheet.cell(row_num, col_num)
            original_header = all_headers.get(col_num, f"Col_{col_num}")
            # Normalize the original header for use as a key (remove spaces, special chars)
            clean_key = f"source_{original_header.replace(' ', '_').replace('(', '').replace(')', '').lower()}"
            all_fields[clean_key] = _clean_cell_value(cell.value)
        
        # Target fields - use original column names  
        for field_type, col_num in column_mapping.target_columns.items():
            cell = worksheet.cell(row_num, col_num)
            original_header = all_headers.get(col_num, f"Col_{col_num}")
            # Normalize the original header for use as a key (remove spaces, special chars)
            clean_key = f"target_{original_header.replace(' ', '_').replace('(', '').replace(')', '').lower()}"
            all_fields[clean_key] = _clean_cell_value(cell.value)
        
        mapping.all_fields = all_fields
        
        # Generate unique ID now that all fields are populated
        mapping.unique_id = mapping.generate_unique_id()
        
        # Check if row has any meaningful data
        if not _has_meaningful_data(all_fields):
            return None
            
        return mapping
        
    except Exception as e:
        logger.warning(f"Error extracting mapping from row {row_num}: {e}")
        return None


def _clean_cell_value(value: Any) -> str:
    """
    Clean and normalize a cell value.
    
    Args:
        value: Raw cell value from openpyxl
        
    Returns:
        Cleaned string value
    """
    if value is None:
        return ""
    
    cleaned = str(value)
    
    if TRIM_WHITESPACE:
        cleaned = cleaned.strip()
    
    return cleaned


def _has_meaningful_data(all_fields: Dict[str, str]) -> bool:
    """
    Check if a mapping row has meaningful data.
    
    Args:
        all_fields: Dictionary of all field values
        
    Returns:
        True if row has sufficient data to be considered a valid mapping
    """
    if IGNORE_EMPTY_CELLS:
        non_empty_count = sum(1 for value in all_fields.values() if value.strip())
        return non_empty_count >= MIN_MAPPING_FIELDS
    
    return len(all_fields) >= MIN_MAPPING_FIELDS


def is_valid_mapping_tab(worksheet: Worksheet) -> bool:
    """
    Check if a worksheet tab has the proper structure for mapping analysis.
    
    Args:
        worksheet: The openpyxl worksheet to validate
        
    Returns:
        True if the tab has valid mapping structure, False otherwise
    """
    try:
        # Check if tab has minimum required rows
        if worksheet.max_row < DATA_START_ROW:
            logger.debug(f"Tab '{worksheet.title}' skipped: insufficient rows ({worksheet.max_row})")
            return False
        
        # Check if tab has minimum required columns
        if worksheet.max_column < 3:
            logger.debug(f"Tab '{worksheet.title}' skipped: insufficient columns ({worksheet.max_column})")
            return False
        
        # Check for proper header structure in first 8 rows
        header_content_found = False
        for row in range(1, 9):  # Rows 1-8 should contain header information
            for col in range(1, min(5, worksheet.max_column + 1)):  # Check first few columns
                cell = worksheet.cell(row, col)
                if cell.value and str(cell.value).strip():
                    # Look for typical header patterns
                    cell_text = str(cell.value).lower().strip()
                    if any(keyword in cell_text for keyword in [
                        'canonical', 'source', 'target', 'system', 'entity', 'field',
                        'description', 'data entity', 'system of record'
                    ]):
                        header_content_found = True
                        break
            if header_content_found:
                break
        
        if not header_content_found:
            logger.debug(f"Tab '{worksheet.title}' skipped: no valid header structure found")
            return False
        
        # Check if row 9 contains system names
        system_names_found = False
        for col in range(1, min(SYSTEM_NAME_MAX_SEARCH_COLUMNS + 1, worksheet.max_column + 1)):
            cell = worksheet.cell(SYSTEM_NAMES_ROW, col)
            if cell.value and str(cell.value).strip() and len(str(cell.value).strip()) >= SYSTEM_NAME_MIN_LENGTH:
                cell_text = str(cell.value).strip()
                # Skip JSON-like content or obvious non-system names
                if not any(char in cell_text for char in ['{', '}', '":', '[', ']']):
                    system_names_found = True
                    break
        
        if not system_names_found:
            logger.debug(f"Tab '{worksheet.title}' skipped: no valid system names in row 9")
            return False
        
        # Check if row 10 contains column headers (must have both source and target sections)
        source_headers_found = False
        target_headers_found = False
        header_positions = []
        
        for col in range(1, min(25, worksheet.max_column + 1)):  # Check more columns for both sections
            cell = worksheet.cell(HEADERS_ROW, col)
            if cell.value and str(cell.value).strip():
                cell_text = str(cell.value).lower().strip()
                if any(keyword in cell_text for keyword in [
                    'canonical', 'field', 'description', 'type', 'length',
                    'format', 'mandatory', 'notes', 'enum', 'entity'
                ]):
                    header_positions.append(col)
        
        # For a valid mapping tab, we need at least 4 relevant headers (suggesting source + target sections)
        # And we need headers spread across the row (not just in first few columns)
        if len(header_positions) >= 4:
            # Check if headers are distributed (suggesting source and target sections)
            min_pos = min(header_positions)
            max_pos = max(header_positions)
            if max_pos - min_pos >= 8:  # Headers should span at least 8 columns for source+target
                source_headers_found = True
                target_headers_found = True
        
        if not (source_headers_found and target_headers_found):
            logger.debug(f"Tab '{worksheet.title}' skipped: insufficient mapping headers in row 10 (found {len(header_positions)} headers)")
            return False
        
        # Check if there's meaningful data starting from row 11
        data_found = False
        for row in range(DATA_START_ROW, min(DATA_START_ROW + 5, worksheet.max_row + 1)):  # Check first 5 data rows
            row_has_data = False
            for col in range(1, min(10, worksheet.max_column + 1)):  # Check first 10 columns
                cell = worksheet.cell(row, col)
                if cell.value and str(cell.value).strip():
                    cell_text = str(cell.value).strip()
                    # Skip JSON-like content
                    if not any(char in cell_text for char in ['{', '}', '":']):
                        row_has_data = True
                        break
            if row_has_data:
                data_found = True
                break
        
        if not data_found:
            logger.debug(f"Tab '{worksheet.title}' skipped: no meaningful data found from row 11")
            return False
        
        return True
        
    except Exception as e:
        logger.warning(f"Error validating tab '{worksheet.title}': {e}")
        return False


def is_hidden_worksheet(worksheet: Worksheet) -> bool:
    """
    Check if a worksheet is hidden.
    
    Args:
        worksheet: The openpyxl worksheet to check
        
    Returns:
        True if the worksheet is hidden, False otherwise
    """
    try:
        # In openpyxl, sheet_state can be 'visible', 'hidden', or 'veryHidden'
        return worksheet.sheet_state != 'visible'
    except Exception as e:
        logger.warning(f"Error checking if worksheet '{worksheet.title}' is hidden: {e}")
        return False


def analyze_worksheet(worksheet: Worksheet) -> TabAnalysis:
    """
    Perform complete analysis of a single worksheet.
    
    Args:
        worksheet: The openpyxl worksheet to analyze
        
    Returns:
        TabAnalysis object with complete analysis results
    """
    analysis = TabAnalysis()
    
    try:
        # Check if tab is hidden and should be skipped
        if is_hidden_worksheet(worksheet):
            # Import current config values to handle runtime changes
            from config import SKIP_HIDDEN_TABS, PROCESS_HIDDEN_TABS
            if SKIP_HIDDEN_TABS and not PROCESS_HIDDEN_TABS:
                analysis.metadata.tab_name = worksheet.title
                analysis.add_error(f"Tab '{worksheet.title}' skipped - worksheet is hidden")
                logger.info(f"Skipping hidden tab '{worksheet.title}'")
                return analysis
            else:
                logger.info(f"Processing hidden tab '{worksheet.title}' (PROCESS_HIDDEN_TABS=True)")
        
        # First validate if this tab should be analyzed
        if not is_valid_mapping_tab(worksheet):
            analysis.metadata.tab_name = worksheet.title
            analysis.add_error(f"Tab '{worksheet.title}' skipped - does not contain valid mapping structure")
            logger.info(f"Skipping tab '{worksheet.title}' - invalid structure")
            return analysis
        
        # Extract metadata
        analysis.metadata = extract_tab_metadata(worksheet)
        logger.info(f"Analyzing tab: {analysis.metadata.tab_name}")
        
        # Identify column structure
        analysis.column_mapping = identify_column_structure(worksheet, analysis.metadata)
        
        # Extract mappings
        analysis.mappings = extract_mappings_from_tab(worksheet, analysis.metadata, analysis.column_mapping)
        
        logger.info(f"Completed analysis of tab '{analysis.metadata.tab_name}': "
                   f"{analysis.mapping_count} mappings found")
        
    except Exception as e:
        error_msg = f"Failed to analyze worksheet '{worksheet.title}': {e}"
        logger.error(error_msg)
        analysis.add_error(error_msg)
    
    return analysis


def analyze_workbook(file_path: str) -> Dict[str, TabAnalysis]:
    """
    Analyze all worksheets in an Excel workbook.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Dictionary mapping tab names to TabAnalysis objects
    """
    results = {}
    
    try:
        logger.info(f"Loading workbook: {file_path}")
        workbook = load_workbook(file_path, data_only=True)
        
        for worksheet in workbook.worksheets:
            analysis = analyze_worksheet(worksheet)
            results[worksheet.title] = analysis
        
        logger.info(f"Completed analysis of workbook '{file_path}': {len(results)} tabs processed")
        
    except Exception as e:
        error_msg = f"Failed to load or analyze workbook '{file_path}': {e}"
        logger.error(error_msg)
        # Return empty results with error information
        error_analysis = TabAnalysis()
        error_analysis.add_error(error_msg)
        results["ERROR"] = error_analysis
    
    return results