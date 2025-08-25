#!/usr/bin/env python3
"""
Debug script for source/target section detection

This script examines how columns are being assigned to source vs target sections
to understand why the Comments column isn't being detected as a target column.
"""

import openpyxl
from config import HEADERS_ROW, MAX_EMPTY_COLUMNS_BETWEEN_SECTIONS
from excel_analyzer import normalize_column_name

def debug_section_detection():
    """Debug source/target section detection."""
    print("Source/Target Section Detection Debug")
    print("="*50)
    
    # Load the workbook
    wb = openpyxl.load_workbook("STTM.xlsx")
    target_sheet = "NetSuiteVendorRequestResponsOTV"
    
    if target_sheet not in wb.sheetnames:
        print(f"ERROR: Sheet '{target_sheet}' not found!")
        return
    
    worksheet = wb[target_sheet]
    print(f"Analyzing sheet: {target_sheet}")
    
    # Extract headers and metadata (simulate detect_tab_metadata)
    print(f"\nExtracting headers from row {HEADERS_ROW}:")
    headers = {}
    max_col = min(worksheet.max_column, 30)  # Check first 30 columns
    
    for col_num in range(1, max_col + 1):
        cell_value = worksheet.cell(HEADERS_ROW, col_num).value
        if cell_value and str(cell_value).strip():
            headers[col_num] = str(cell_value).strip()
            col_letter = openpyxl.utils.get_column_letter(col_num)
            print(f"  {col_letter}{HEADERS_ROW} (col {col_num}): '{cell_value}'")
    
    # Simulate target system detection (col 14 = NetSuite Vendor Work Item)
    target_system_column = 14  # Column N
    print(f"\nTarget system column: {target_system_column} (NetSuite Vendor Work Item)")
    print(f"Max empty columns between sections: {MAX_EMPTY_COLUMNS_BETWEEN_SECTIONS}")
    
    # Normalize headers and find duplicates
    print(f"\nNormalizing headers:")
    header_counts = {}
    for col_num, header in headers.items():
        normalized = normalize_column_name(header)
        col_letter = openpyxl.utils.get_column_letter(col_num)
        print(f"  {col_letter} '{header}' -> '{normalized}'")
        
        if normalized:
            if normalized not in header_counts:
                header_counts[normalized] = []
            header_counts[normalized].append(col_num)
    
    # Show duplicate detection
    print(f"\nDuplicate column types:")
    source_columns = []
    target_columns = []
    assigned_target = set()
    
    for normalized_name, col_list in header_counts.items():
        if len(col_list) >= 2:
            col_list.sort()
            print(f"  {normalized_name}: columns {col_list}")
            print(f"    -> Source: {col_list[0]}, Target: {col_list[1:]}")
            source_columns.extend(col_list[:1])
            target_columns.extend(col_list[1:])
            assigned_target.update(col_list[1:])
    
    # Boundary-based assignment for remaining columns
    print(f"\nBoundary-based assignment (boundary = col {target_system_column}):")
    for col_num, header in headers.items():
        if col_num not in assigned_target and col_num not in source_columns:
            col_letter = openpyxl.utils.get_column_letter(col_num)
            
            if col_num < target_system_column:
                assignment = "SOURCE (before boundary)"
                source_columns.append(col_num)
            else:
                # Check gap
                gap_size = count_empty_columns_before(col_num, headers, target_system_column)
                print(f"    {col_letter} '{header}': gap_size = {gap_size}")
                
                if gap_size <= MAX_EMPTY_COLUMNS_BETWEEN_SECTIONS:
                    assignment = "TARGET (gap ok)"
                    target_columns.append(col_num)
                else:
                    assignment = f"SOURCE (gap too large: {gap_size})"
                    source_columns.append(col_num)
            
            print(f"    {col_letter} '{header}' -> {assignment}")
    
    # Final results
    print(f"\nFinal column assignments:")
    print(f"Source columns ({len(source_columns)}): {sorted(source_columns)}")
    print(f"Target columns ({len(target_columns)}): {sorted(target_columns)}")
    
    # Check specifically for Comments column
    comments_col = 16  # Column P
    if comments_col in target_columns:
        print(f"\n✅ Comments column (P/{comments_col}) correctly assigned to TARGET")
    elif comments_col in source_columns:
        print(f"\n❌ Comments column (P/{comments_col}) incorrectly assigned to SOURCE")
    else:
        print(f"\n❌ Comments column (P/{comments_col}) not assigned to either section!")
    
    wb.close()


def count_empty_columns_before(col_num: int, headers: dict, start_col: int) -> int:
    """Count empty columns between start_col and col_num."""
    if col_num <= start_col:
        return 0
    
    empty_count = 0
    for check_col in range(start_col + 1, col_num):
        if check_col not in headers or not headers.get(check_col, "").strip():
            empty_count += 1
    
    return empty_count


if __name__ == "__main__":
    debug_section_detection()