"""
Test script for Phase 2: Comparison Engine

This script tests the comparison functionality by creating test scenarios
and validating the comparison results.
"""

import logging
import shutil
from pathlib import Path
from comparator import compare_workbooks, create_test_comparison
from utils import validate_file_path, get_comparison_statistics
import openpyxl

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

def create_test_excel_file(original_file: str, test_file: str, modifications: dict):
    """
    Create a test Excel file by modifying the original file.
    
    Args:
        original_file: Path to original Excel file
        test_file: Path where test file will be created
        modifications: Dictionary describing modifications to make
    """
    print(f"Creating test file: {test_file}")
    
    # Copy original file
    shutil.copy2(original_file, test_file)
    
    # Open and modify the test file
    wb = openpyxl.load_workbook(test_file)
    
    for mod_type, details in modifications.items():
        if mod_type == "delete_tab":
            tab_name = details
            if tab_name in wb.sheetnames:
                del wb[tab_name]
                print(f"  Deleted tab: {tab_name}")
        
        elif mod_type == "add_tab":
            tab_name = details["name"]
            source_tab = details.get("copy_from")
            if source_tab and source_tab in wb.sheetnames:
                # Copy existing tab
                source_sheet = wb[source_tab]
                new_sheet = wb.copy_worksheet(source_sheet)
                new_sheet.title = tab_name
                print(f"  Added tab: {tab_name} (copied from {source_tab})")
            else:
                # Create new empty tab
                wb.create_sheet(tab_name)
                print(f"  Added empty tab: {tab_name}")
        
        elif mod_type == "modify_mapping":
            tab_name = details["tab"]
            row = details["row"]
            col = details["col"]
            new_value = details["value"]
            
            if tab_name in wb.sheetnames:
                sheet = wb[tab_name]
                old_value = sheet.cell(row, col).value
                sheet.cell(row, col).value = new_value
                print(f"  Modified {tab_name}[{row},{col}]: '{old_value}' -> '{new_value}'")
        
        elif mod_type == "add_mapping":
            tab_name = details["tab"]
            row = details["row"]
            mapping_data = details["data"]
            
            if tab_name in wb.sheetnames:
                sheet = wb[tab_name]
                for col, value in enumerate(mapping_data, 1):
                    sheet.cell(row, col).value = value
                print(f"  Added mapping to {tab_name} at row {row}")
    
    # Save modified file
    wb.save(test_file)
    wb.close()


def test_basic_comparison():
    """Test basic comparison functionality with the same file."""
    print("\n" + "="*60)
    print("TEST 1: Basic Comparison (Same File)")
    print("="*60)
    
    original_file = "STTM.xlsx"
    
    # Validate file exists
    is_valid, error = validate_file_path(original_file)
    if not is_valid:
        print(f"ERROR: {error}")
        return False
    
    # Compare file with itself
    result = create_test_comparison(original_file, original_file, output_summary=True)
    
    # Validate results
    if result.has_errors:
        print("FAILED: Comparison had errors")
        return False
    
    # Should have no changes when comparing file with itself
    total_changes = (result.summary.tabs_added + result.summary.tabs_deleted + 
                    result.summary.tabs_modified)
    
    if total_changes == 0:
        print("[PASS] PASSED: No changes detected when comparing file with itself")
        return True
    else:
        print(f"[FAIL] FAILED: Expected 0 changes, but found {total_changes}")
        return False


def test_tab_deletion():
    """Test detection of deleted tabs."""
    print("\n" + "="*60)
    print("TEST 2: Tab Deletion Detection")
    print("="*60)
    
    original_file = "STTM.xlsx"
    test_file = "test_deleted_tab.xlsx"
    
    try:
        # Create test file with deleted tab
        modifications = {
            "delete_tab": "NetSuiteVendorRequestResponsOTV"  # Delete first valid tab
        }
        create_test_excel_file(original_file, test_file, modifications)
        
        # Compare files
        result = create_test_comparison(original_file, test_file, output_summary=True)
        
        if result.has_errors:
            print("FAILED: Comparison had errors")
            return False
        
        # Should detect 1 deleted tab
        if result.summary.tabs_deleted == 1:
            print("[PASS] PASSED: Correctly detected tab deletion")
            return True
        else:
            print(f"[FAIL] FAILED: Expected 1 deleted tab, found {result.summary.tabs_deleted}")
            return False
    
    finally:
        # Clean up test file
        if Path(test_file).exists():
            Path(test_file).unlink()


def test_tab_addition():
    """Test detection of added tabs."""
    print("\n" + "="*60)
    print("TEST 3: Tab Addition Detection")
    print("="*60)
    
    original_file = "STTM.xlsx"
    test_file = "test_added_tab.xlsx"
    
    try:
        # Create test file with added tab (copy of existing tab)
        modifications = {
            "add_tab": {
                "name": "NewTestTab",
                "copy_from": "NetSuiteVendorRequestResponsOTV"
            }
        }
        create_test_excel_file(original_file, test_file, modifications)
        
        # Compare files
        result = create_test_comparison(original_file, test_file, output_summary=True)
        
        if result.has_errors:
            print("FAILED: Comparison had errors")
            return False
        
        # Should detect 1 added tab
        if result.summary.tabs_added == 1:
            print("[PASS] PASSED: Correctly detected tab addition")
            return True
        else:
            print(f"[FAIL] FAILED: Expected 1 added tab, found {result.summary.tabs_added}")
            return False
    
    finally:
        # Clean up test file
        if Path(test_file).exists():
            Path(test_file).unlink()


def test_mapping_modification():
    """Test detection of modified mappings."""
    print("\n" + "="*60)
    print("TEST 4: Mapping Modification Detection")
    print("="*60)
    
    original_file = "STTM.xlsx"
    test_file = "test_modified_mapping.xlsx"
    
    try:
        # Create test file with modified mapping
        modifications = {
            "modify_mapping": {
                "tab": "NetSuiteVendorRequestResponsOTV",
                "row": 17,  # Row with mapping data (first actual mapping)
                "col": 4,   # Description column
                "value": "MODIFIED TEST DESCRIPTION"
            }
        }
        create_test_excel_file(original_file, test_file, modifications)
        
        # Compare files
        result = create_test_comparison(original_file, test_file, output_summary=True)
        
        if result.has_errors:
            print("FAILED: Comparison had errors")
            return False
        
        # Should detect at least 1 modified mapping
        if result.summary.total_mappings_modified >= 1:
            print("[PASS] PASSED: Correctly detected mapping modification")
            return True
        else:
            print(f"[FAIL] FAILED: Expected >=1 modified mapping, found {result.summary.total_mappings_modified}")
            return False
    
    finally:
        # Clean up test file
        if Path(test_file).exists():
            Path(test_file).unlink()


def test_complex_changes():
    """Test detection of multiple types of changes."""
    print("\n" + "="*60)
    print("TEST 5: Complex Changes Detection")
    print("="*60)
    
    original_file = "STTM.xlsx"
    test_file = "test_complex_changes.xlsx"
    
    try:
        # Create test file with multiple changes
        modifications = {
            "delete_tab": "NetSuiteVendorReqRespAssociate",  # Delete one tab
            "add_tab": {
                "name": "ComplexTestTab", 
                "copy_from": "Vendor Inbound DACH VenProxy"
            },
            "modify_mapping": {
                "tab": "VendorInboundVendorProxytoD365",
                "row": 13,
                "col": 4,  # Description field (non-core)
                "value": "COMPLEX TEST MODIFICATION"
            }
        }
        create_test_excel_file(original_file, test_file, modifications)
        
        # Compare files
        result = create_test_comparison(original_file, test_file, output_summary=True)
        
        if result.has_errors:
            print("FAILED: Comparison had errors")
            return False
        
        # Should detect multiple types of changes
        has_deleted = result.summary.tabs_deleted >= 1
        has_added = result.summary.tabs_added >= 1
        has_modified = result.summary.total_mappings_modified >= 1
        
        if has_deleted and has_added and has_modified:
            print("[PASS] PASSED: Correctly detected multiple types of changes")
            return True
        else:
            print(f"[FAIL] FAILED: Missing change detection - Deleted:{has_deleted}, Added:{has_added}, Modified:{has_modified}")
            return False
    
    finally:
        # Clean up test file
        if Path(test_file).exists():
            Path(test_file).unlink()


def test_detailed_statistics():
    """Test detailed statistics generation."""
    print("\n" + "="*60)
    print("TEST 6: Detailed Statistics")
    print("="*60)
    
    original_file = "STTM.xlsx"
    test_file = "test_statistics.xlsx"
    
    try:
        # Create test file with known changes
        modifications = {
            "modify_mapping": {
                "tab": "VendorInboundVendorProxytoD365",
                "row": 13,  # Row with existing description
                "col": 4,   # Description column  
                "value": "STATISTICS TEST VALUE"
            }
        }
        create_test_excel_file(original_file, test_file, modifications)
        
        # Compare files
        result = compare_workbooks(original_file, test_file)
        
        if result.has_errors:
            print("FAILED: Comparison had errors")
            return False
        
        # Get detailed statistics
        stats = get_comparison_statistics(result)
        
        print(f"Detailed Statistics:")
        print(f"  Total tabs compared: {stats['summary']['total_tabs_compared']}")
        print(f"  Tabs with changes: {stats['summary']['tabs_with_changes']}")
        print(f"  Total changes: {stats['summary']['total_changes']}")
        print(f"  Mapping changes: {stats['mapping_changes']}")
        
        if stats['tab_changes']:
            print(f"  Changed tabs:")
            for tab_name, changes in stats['tab_changes'].items():
                print(f"    {tab_name}: {changes}")
        
        if stats['summary']['tabs_with_changes'] > 0:
            print("[PASS] PASSED: Statistics generated successfully")
            return True
        else:
            print("[FAIL] FAILED: No changes detected in statistics")
            return False
    
    finally:
        # Clean up test file
        if Path(test_file).exists():
            Path(test_file).unlink()


def run_all_tests():
    """Run all Phase 2 tests."""
    print("Starting Phase 2 Comparison Engine Tests")
    print("="*60)
    
    tests = [
        ("Basic Comparison", test_basic_comparison),
        ("Tab Deletion", test_tab_deletion),
        ("Tab Addition", test_tab_addition),
        ("Mapping Modification", test_mapping_modification),
        ("Complex Changes", test_complex_changes),
        ("Detailed Statistics", test_detailed_statistics)
    ]
    
    results = []
    
    for test_name, test_func in tests:
        try:
            result = test_func()
            results.append((test_name, result))
        except Exception as e:
            print(f"[FAIL] ERROR in {test_name}: {e}")
            results.append((test_name, False))
    
    # Summary
    print("\n" + "="*60)
    print("TEST RESULTS SUMMARY")
    print("="*60)
    
    passed = sum(1 for _, result in results if result)
    total = len(results)
    
    for test_name, result in results:
        status = "PASSED" if result else "FAILED"
        print(f"{status} - {test_name}")
    
    print(f"\nOverall: {passed}/{total} tests passed ({passed/total*100:.1f}%)")
    
    if passed == total:
        print("All Phase 2 tests passed! Comparison engine is working correctly.")
        return True
    else:
        print("Some tests failed. Please review the comparison logic.")
        return False


def test_with_custom_files(file1: str, file2: str):
    """Test comparison with custom file paths."""
    # Validate files first
    print("Validating input files...")
    for i, file_path in enumerate([file1, file2], 1):
        is_valid, error = validate_file_path(file_path)
        if not is_valid:
            print(f"ERROR: File {i} validation failed: {error}")
            return False
        else:
            print(f"  File {i}: {file_path} - OK")
    
    print("\nStarting comparison...")
    
    # Run comparison
    result = create_test_comparison(file1, file2, output_summary=True)
    
    if result.has_errors:
        print("\nCOMPARISSON ERRORS:")
        for error in result.errors:
            print(f"  - {error}")
        return False
    else:
        print("\nComparison completed successfully!")
        
        # Show additional details
        print(f"\nDETAILED RESULTS:")
        print(f"  Files compared: 2")
        print(f"  Tabs in file 1: {result.summary.total_tabs_v1}")
        print(f"  Tabs in file 2: {result.summary.total_tabs_v2}")
        print(f"  Total mappings in file 1: {result.summary.total_mappings_v1}")
        print(f"  Total mappings in file 2: {result.summary.total_mappings_v2}")
        
        # Show change details if any
        total_changes = (result.summary.tabs_added + result.summary.tabs_deleted + 
                        result.summary.tabs_modified + result.summary.total_mappings_added +
                        result.summary.total_mappings_deleted + result.summary.total_mappings_modified)
        
        if total_changes > 0:
            print(f"\nCHANGES DETECTED:")
            if result.summary.tabs_added > 0:
                print(f"  Tabs added: {result.summary.tabs_added}")
            if result.summary.tabs_deleted > 0:
                print(f"  Tabs deleted: {result.summary.tabs_deleted}")
            if result.summary.tabs_modified > 0:
                print(f"  Tabs modified: {result.summary.tabs_modified}")
            if result.summary.total_mappings_added > 0:
                print(f"  Mappings added: {result.summary.total_mappings_added}")
            if result.summary.total_mappings_deleted > 0:
                print(f"  Mappings deleted: {result.summary.total_mappings_deleted}")
            if result.summary.total_mappings_modified > 0:
                print(f"  Mappings modified: {result.summary.total_mappings_modified}")
                
            print(f"\nCHANGED TABS BREAKDOWN:")
            for tab_name, comparison in result.tab_comparisons.items():
                if comparison.has_changes:
                    changes = comparison.change_summary
                    print(f"  {tab_name}: +{changes['added']} -{changes['deleted']} ~{changes['modified']}")
        else:
            print(f"\nNO CHANGES DETECTED: Files are identical")
        
        return True


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) == 3:
        # Custom file comparison - enhanced version
        file1, file2 = sys.argv[1], sys.argv[2]
        print(f"Excel Comparison Tool - Phase 2 Test")
        print("=" * 60)
        print(f"Comparing: {file1}")
        print(f"     vs:   {file2}")
        print("=" * 60)
        
        success = test_with_custom_files(file1, file2)
        
        if success:
            print("\n" + "=" * 60)
            print("COMPARISON COMPLETED SUCCESSFULLY")
            print("=" * 60)
        else:
            print("\n" + "=" * 60) 
            print("COMPARISON FAILED - CHECK ERRORS ABOVE")
            print("=" * 60)
            sys.exit(1)
            
    elif len(sys.argv) > 3:
        print("Usage: python test_phase2.py [file1.xlsx] [file2.xlsx]")
        print("  With no arguments: Run all standard tests")
        print("  With 2 arguments: Compare two specific Excel files")
        sys.exit(1)
    else:
        # Run all standard tests
        run_all_tests()