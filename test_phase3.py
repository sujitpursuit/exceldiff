"""
Test script for Phase 3: HTML Report Generator

This script tests the HTML report generation functionality by creating various
test scenarios and validating the HTML output.
"""

import logging
import shutil
from pathlib import Path
from datetime import datetime
import openpyxl

from comparator import compare_workbooks
from report_generator import HTMLReportGenerator, generate_html_report
from utils import validate_file_path

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)


def create_test_report_scenarios():
    """Create various test scenarios for report generation."""
    scenarios = []
    
    # Scenario 1: Basic comparison (same file)
    scenarios.append({
        'name': 'Basic Comparison (Same File)',
        'file1': 'STTM.xlsx',
        'file2': 'STTM.xlsx',
        'output': 'test_report_basic.html',
        'description': 'Test report generation with identical files'
    })
    
    # Scenario 2: Files with modifications
    scenarios.append({
        'name': 'Modified Mappings',
        'file1': 'STTM.xlsx',
        'file2': 'test_report_modified.xlsx',
        'output': 'test_report_modified.html',
        'description': 'Test report generation with field modifications',
        'create_file2': lambda: create_modified_test_file('STTM.xlsx', 'test_report_modified.xlsx')
    })
    
    # Scenario 3: Complex changes (add, delete, modify)
    scenarios.append({
        'name': 'Complex Changes',
        'file1': 'STTM.xlsx',
        'file2': 'test_report_complex.xlsx',
        'output': 'test_report_complex.html',
        'description': 'Test report generation with mixed change types',
        'create_file2': lambda: create_complex_test_file('STTM.xlsx', 'test_report_complex.xlsx')
    })
    
    return scenarios


def create_modified_test_file(original_file: str, test_file: str):
    """Create a test file with modifications."""
    print(f"Creating modified test file: {test_file}")
    
    # Copy original file
    shutil.copy2(original_file, test_file)
    
    # Open and modify the test file
    wb = openpyxl.load_workbook(test_file)
    
    # Modify a few mappings in NetSuiteVendorRequestResponsOTV
    if 'NetSuiteVendorRequestResponsOTV' in wb.sheetnames:
        sheet = wb['NetSuiteVendorRequestResponsOTV']
        
        # Modify description in row 17
        old_desc = sheet.cell(17, 4).value
        sheet.cell(17, 4).value = 'MODIFIED: Updated correlation ID format for Phase 3 testing'
        print(f"  Modified description: {old_desc} -> Updated format")
        
        # Modify type in row 18
        sheet.cell(18, 5).value = 'Text'  # Change type to Text
        print(f"  Modified field type to Text")
    
    wb.save(test_file)
    wb.close()


def create_complex_test_file(original_file: str, test_file: str):
    """Create a test file with complex changes (add, delete, modify)."""
    print(f"Creating complex test file: {test_file}")
    
    # Copy original file
    shutil.copy2(original_file, test_file)
    
    # Open and modify the test file
    wb = openpyxl.load_workbook(test_file)
    
    # Delete a tab to simulate tab deletion
    if 'NetSuiteVendorReqRespAssociate' in wb.sheetnames:
        del wb['NetSuiteVendorReqRespAssociate']
        print("  Deleted tab: NetSuiteVendorReqRespAssociate")
    
    # Add a new tab by copying existing one
    if 'NetSuiteVendorRequestResponsOTV' in wb.sheetnames:
        source_sheet = wb['NetSuiteVendorRequestResponsOTV']
        new_sheet = wb.copy_worksheet(source_sheet)
        new_sheet.title = 'NewTestMappingTab'
        print("  Added tab: NewTestMappingTab")
    
    # Modify existing mappings
    if 'VendorInboundVendorProxytoD365' in wb.sheetnames:
        sheet = wb['VendorInboundVendorProxytoD365']
        
        # Modify some descriptions
        sheet.cell(13, 4).value = 'MODIFIED: Complex test scenario change'
        sheet.cell(14, 4).value = 'MODIFIED: Another field description update'
        print("  Modified 2 field descriptions")
    
    wb.save(test_file)
    wb.close()


def test_report_generation():
    """Test basic report generation functionality."""
    print("\n" + "="*60)
    print("TEST 1: Basic Report Generation")
    print("="*60)
    
    # Use STTM.xlsx vs itself for basic test
    original_file = "STTM.xlsx"
    
    # Validate file exists
    is_valid, error = validate_file_path(original_file)
    if not is_valid:
        print(f"ERROR: {error}")
        return False
    
    try:
        # Compare files
        print("Running comparison...")
        result = compare_workbooks(original_file, original_file)
        
        if result.has_errors:
            print("FAILED: Comparison had errors")
            return False
        
        # Generate HTML report
        print("Generating HTML report...")
        output_path = "reports/test_reports/test_basic_report.html"
        success = generate_html_report(result, output_path, "Basic Test Report")
        
        if success:
            print(f"[PASS] PASSED: Report generated successfully at {output_path}")
            
            # Validate HTML file was created and has content
            if Path(output_path).exists() and Path(output_path).stat().st_size > 0:
                print(f"  File size: {Path(output_path).stat().st_size} bytes")
                return True
            else:
                print("[FAIL] FAILED: Report file is empty or not created")
                return False
        else:
            print("[FAIL] FAILED: Report generation failed")
            return False
            
    except Exception as e:
        print(f"[FAIL] ERROR: {e}")
        return False


def test_report_with_changes():
    """Test report generation with actual changes."""
    print("\n" + "="*60)
    print("TEST 2: Report Generation with Changes")
    print("="*60)
    
    original_file = "STTM.xlsx"
    test_file = "test_changes_report.xlsx"
    
    try:
        # Create test file with modifications
        create_modified_test_file(original_file, test_file)
        
        # Compare files
        print("Running comparison...")
        result = compare_workbooks(original_file, test_file)
        
        if result.has_errors:
            print("FAILED: Comparison had errors")
            return False
        
        # Check that we have some changes
        total_changes = (result.summary.tabs_added + result.summary.tabs_deleted + 
                        result.summary.tabs_modified + result.summary.total_mappings_added +
                        result.summary.total_mappings_deleted + result.summary.total_mappings_modified)
        
        print(f"Changes detected: {total_changes}")
        
        # Generate HTML report
        print("Generating HTML report with changes...")
        output_path = "reports/test_reports/test_changes_report.html"
        success = generate_html_report(result, output_path, "Changes Test Report")
        
        if success:
            print(f"[PASS] PASSED: Report with changes generated successfully at {output_path}")
            return True
        else:
            print("[FAIL] FAILED: Report generation failed")
            return False
            
    except Exception as e:
        print(f"[FAIL] ERROR: {e}")
        return False
    finally:
        # Clean up test file
        if Path(test_file).exists():
            Path(test_file).unlink()


def test_complex_report():
    """Test report generation with complex changes."""
    print("\n" + "="*60)
    print("TEST 3: Complex Changes Report")
    print("="*60)
    
    original_file = "STTM.xlsx"
    test_file = "test_complex_report.xlsx"
    
    try:
        # Create test file with complex changes
        create_complex_test_file(original_file, test_file)
        
        # Compare files
        print("Running comparison...")
        result = compare_workbooks(original_file, test_file)
        
        if result.has_errors:
            print("FAILED: Comparison had errors")
            return False
        
        # Show change summary
        summary = result.summary
        print(f"Tabs: +{summary.tabs_added} -{summary.tabs_deleted} ~{summary.tabs_modified}")
        print(f"Mappings: +{summary.total_mappings_added} -{summary.total_mappings_deleted} ~{summary.total_mappings_modified}")
        
        # Generate HTML report
        print("Generating complex changes report...")
        output_path = "reports/test_reports/test_complex_report.html"
        success = generate_html_report(result, output_path, "Complex Changes Report")
        
        if success:
            print(f"[PASS] PASSED: Complex report generated successfully at {output_path}")
            return True
        else:
            print("[FAIL] FAILED: Report generation failed")
            return False
            
    except Exception as e:
        print(f"[FAIL] ERROR: {e}")
        return False
    finally:
        # Clean up test file
        if Path(test_file).exists():
            Path(test_file).unlink()


def test_report_class_functionality():
    """Test the HTMLReportGenerator class directly."""
    print("\n" + "="*60)
    print("TEST 4: HTMLReportGenerator Class")
    print("="*60)
    
    try:
        # Create generator instance
        generator = HTMLReportGenerator()
        print("HTMLReportGenerator instance created successfully")
        
        # Run basic comparison
        result = compare_workbooks("STTM.xlsx", "STTM.xlsx")
        
        # Test different methods
        output_path = "reports/test_reports/test_class_report.html"
        success = generator.generate_report(result, output_path, "Class Test Report")
        
        if success and Path(output_path).exists():
            print("[PASS] PASSED: HTMLReportGenerator class works correctly")
            
            # Read a bit of the generated HTML to verify structure
            with open(output_path, 'r', encoding='utf-8') as f:
                html_content = f.read(500)  # Read first 500 chars
                
            if '<!DOCTYPE html>' in html_content and '<title>' in html_content:
                print("  HTML structure verification passed")
                return True
            else:
                print("[FAIL] FAILED: HTML structure validation failed")
                return False
        else:
            print("[FAIL] FAILED: HTMLReportGenerator class test failed")
            return False
            
    except Exception as e:
        print(f"[FAIL] ERROR: {e}")
        return False


def test_html_content_validation():
    """Test HTML content validation and structure."""
    print("\n" + "="*60)
    print("TEST 5: HTML Content Validation")
    print("="*60)
    
    try:
        # Generate a report
        result = compare_workbooks("STTM.xlsx", "STTM.xlsx")
        output_path = "reports/test_reports/test_validation_report.html"
        success = generate_html_report(result, output_path)
        
        if not success:
            print("[FAIL] FAILED: Could not generate report for validation")
            return False
        
        # Read and validate HTML content
        with open(output_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
        
        # Check for required HTML elements
        required_elements = [
            '<!DOCTYPE html>',
            '<html lang="en">',
            '<head>',
            '<title>',
            '<style>',
            '<body>',
            'Executive Summary',
            'Detailed Changes',
            'Technical Details',
            '</html>'
        ]
        
        missing_elements = []
        for element in required_elements:
            if element not in html_content:
                missing_elements.append(element)
        
        if missing_elements:
            print(f"[FAIL] FAILED: Missing HTML elements: {missing_elements}")
            return False
        
        # Check for CSS styles
        css_classes = [
            '.header',
            '.summary-section',
            '.changes-section',
            '.mapping-table'
        ]
        
        missing_css = []
        for css_class in css_classes:
            if css_class not in html_content:
                missing_css.append(css_class)
        
        if missing_css:
            print(f"[FAIL] FAILED: Missing CSS classes: {missing_css}")
            return False
        
        print("[PASS] PASSED: HTML content validation successful")
        print(f"  Report size: {len(html_content)} characters")
        print(f"  Contains all required elements and styles")
        return True
        
    except Exception as e:
        print(f"[FAIL] ERROR: {e}")
        return False


def run_all_tests():
    """Run all Phase 3 tests."""
    print("Starting Phase 3 HTML Report Generator Tests")
    print("="*60)
    
    tests = [
        ("Basic Report Generation", test_report_generation),
        ("Report with Changes", test_report_with_changes),
        ("Complex Changes Report", test_complex_report),
        ("HTMLReportGenerator Class", test_report_class_functionality),
        ("HTML Content Validation", test_html_content_validation)
    ]
    
    results = []
    
    for test_name, test_func in tests:
        try:
            result = test_func()
            results.append((test_name, result))
        except Exception as e:
            print(f"[FAIL] ERROR in {test_name}: {e}")
            results.append((test_name, False))
    
    # Summary
    print("\n" + "="*60)
    print("TEST RESULTS SUMMARY")
    print("="*60)
    
    passed = sum(1 for _, result in results if result)
    total = len(results)
    
    for test_name, result in results:
        status = "PASSED" if result else "FAILED"
        print(f"{status} - {test_name}")
    
    print(f"\nOverall: {passed}/{total} tests passed ({passed/total*100:.1f}%)")
    
    if passed == total:
        print("All Phase 3 tests passed! HTML Report Generator is working correctly.")
        print("\nGenerated test reports:")
        test_files = [
            "reports/test_reports/test_basic_report.html",
            "reports/test_reports/test_changes_report.html", 
            "reports/test_reports/test_complex_report.html",
            "reports/test_reports/test_class_report.html",
            "reports/test_reports/test_validation_report.html"
        ]
        for file_name in test_files:
            if Path(file_name).exists():
                print(f"  - {file_name}")
        return True
    else:
        print("Some tests failed. Please review the HTML report generation logic.")
        return False


def test_custom_scenarios():
    """Test report generation with custom comparison files."""
    print("\n" + "="*60)
    print("CUSTOM TEST: Report Generation with Custom Files")
    print("="*60)
    
    import sys
    
    if len(sys.argv) >= 3:
        file1, file2 = sys.argv[1], sys.argv[2]
        output_file = sys.argv[3] if len(sys.argv) > 3 else "custom_comparison_report.html"
        
        # Validate files
        for file_path in [file1, file2]:
            is_valid, error = validate_file_path(file_path)
            if not is_valid:
                print(f"ERROR: {error}")
                return False
        
        try:
            print(f"Comparing {file1} vs {file2}")
            result = compare_workbooks(file1, file2)
            
            if result.has_errors:
                print("FAILED: Comparison had errors")
                return False
            
            # Ensure output file is in reports directory  
            if not output_file.startswith("reports/"):
                output_file = f"reports/test_reports/{output_file}"
            
            print(f"Generating report: {output_file}")
            success = generate_html_report(result, output_file, f"Comparison: {Path(file1).name} vs {Path(file2).name}")
            
            if success:
                print(f"[PASS] PASSED: Custom report generated successfully")
                print(f"  Output: {output_file}")
                return True
            else:
                print("[FAIL] FAILED: Custom report generation failed")
                return False
                
        except Exception as e:
            print(f"[FAIL] ERROR: {e}")
            return False
    else:
        print("Usage: python test_phase3.py [file1.xlsx] [file2.xlsx] [output.html]")
        print("Run without arguments for standard test suite")
        return True


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) >= 3:
        # Custom file testing
        test_custom_scenarios()
    else:
        # Run all standard tests
        run_all_tests()