#!/usr/bin/env python3
"""
Debug script for Column Header Detection

This script examines what column headers are being detected in the 
NetSuiteVendorRequestResponsOTV tab, specifically looking for the 
"Comments" column in column P.
"""

import openpyxl
from config import HEADERS_ROW

def debug_column_headers():
    """Debug column header detection in NetSuiteVendorRequestResponsOTV tab."""
    print("Column Header Detection Debug")
    print("="*50)
    
    # Load the workbook
    wb = openpyxl.load_workbook("STTM.xlsx")
    target_sheet = "NetSuiteVendorRequestResponsOTV"
    
    if target_sheet not in wb.sheetnames:
        print(f"ERROR: Sheet '{target_sheet}' not found!")
        return
    
    worksheet = wb[target_sheet]
    
    print(f"Analyzing sheet: {target_sheet}")
    print(f"Headers row: {HEADERS_ROW} (Row 10)")
    
    # Show raw headers in row 10
    print(f"\nRaw headers in row {HEADERS_ROW}:")
    max_col = min(worksheet.max_column, 30)  # Check first 30 columns
    
    for col_num in range(1, max_col + 1):
        cell_value = worksheet.cell(HEADERS_ROW, col_num).value
        col_letter = openpyxl.utils.get_column_letter(col_num)
        
        if cell_value:
            print(f"  {col_letter}{HEADERS_ROW} (col {col_num}): '{cell_value}'")
    
    # Check specifically for column P (16)
    print(f"\nSpecific check for column P (16):")
    p_value = worksheet.cell(HEADERS_ROW, 16).value
    print(f"  P{HEADERS_ROW}: '{p_value}'")
    
    # Use the actual analyzer function to see what gets detected
    print(f"\nUsing analyzer functions:")
    try:
        from excel_analyzer import detect_tab_metadata, identify_column_structure
        
        # First get tab metadata (including system names)
        metadata = detect_tab_metadata(worksheet)
        print(f"  Source system: '{metadata.source_system}' (col {metadata.source_system_column})")  
        print(f"  Target system: '{metadata.target_system}' (col {metadata.target_system_column})")
        
        # Now get column mapping
        column_mapping = identify_column_structure(worksheet, metadata)
        
        print(f"\nDetected source columns: {len(column_mapping.source_columns)}")
        for field_type, col_num in column_mapping.source_columns.items():
            col_letter = openpyxl.utils.get_column_letter(col_num)
            header_value = worksheet.cell(HEADERS_ROW, col_num).value
            print(f"  {field_type}: {col_letter}{HEADERS_ROW} (col {col_num}) = '{header_value}'")
        
        print(f"\nDetected target columns: {len(column_mapping.target_columns)}")
        for field_type, col_num in column_mapping.target_columns.items():
            col_letter = openpyxl.utils.get_column_letter(col_num)
            header_value = worksheet.cell(HEADERS_ROW, col_num).value
            print(f"  {field_type}: {col_letter}{HEADERS_ROW} (col {col_num}) = '{header_value}'")
        
        # Check if Comments/description is detected
        if 'description' in column_mapping.target_columns:
            desc_col = column_mapping.target_columns['description']
            desc_letter = openpyxl.utils.get_column_letter(desc_col)
            desc_value = worksheet.cell(HEADERS_ROW, desc_col).value
            print(f"\n✅ Description/Comments field detected:")
            print(f"   Column: {desc_letter}{HEADERS_ROW} (col {desc_col})")
            print(f"   Header: '{desc_value}'")
        else:
            print(f"\n❌ Description/Comments field NOT detected!")
            print(f"   Expected: Column P16 with header 'Comments'")
            
            # Show what's actually in P16
            p16_value = worksheet.cell(16, 16).value  # Row 16, Col P 
            print(f"   Actual P16 value: '{p16_value}'")
        
    except Exception as e:
        print(f"Error calling analyzer functions: {e}")
    
    wb.close()


def check_comments_mapping():
    """Check the column name mapping for comments."""
    print(f"\nColumn Name Mapping Check")
    print("="*30)
    
    from config import COLUMN_NAME_MAPPINGS
    
    if 'description' in COLUMN_NAME_MAPPINGS:
        desc_variants = COLUMN_NAME_MAPPINGS['description']
        print(f"Description field variants: {desc_variants}")
        
        if 'comments' in desc_variants:
            print("✅ 'comments' is mapped to description field")
        else:
            print("❌ 'comments' is NOT mapped to description field")
    else:
        print("❌ No description field mapping found!")


if __name__ == "__main__":
    debug_column_headers()
    check_comments_mapping()